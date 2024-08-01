import pandas as pd
import os
import numpy as np
import pyautogui
from datetime import datetime
import pyautogui
import pyperclip
from datetime import datetime, timedelta
from openpyxl import load_workbook


def aviso_atencao(image_path, image_path2,image_path3,image_path4, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = 'IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    image_path4 = os.path.join(current_dir, caminho_imagem, image_path4)
    attempts = 3
    aviso_ativo = False  # Variável para indicar se o aviso está ativo
    aviso_veiculo = False
    # Tentar encontrar a primeira imagem até 3 vezes
    for attempt in range(attempts):
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                aviso_ativo = True  # Aviso encontrado, definir como ativo
                break
            else:
                print(f"Imagem 1 não encontrada na tentativa {attempt + 1}. Aguardando...")
        except Exception as e:
            print(f"Erro ao tentar encontrar a imagem 1 na tentativa {attempt + 1}: {e}")
        try:
            position4 = pyautogui.locateOnScreen(image_path4, confidence=confidence)
            if position4:
                click_image('ok_marcado.png')
                aviso_veiculo = True  # Aviso encontrado, definir como ativo
                break
        except Exception as e:
            print(f"Erro ao tentar encontrar a imagem 1 na tentativa {attempt + 1}: {e}")    
        pyautogui.sleep(1)
        
    if not aviso_ativo:
        print("Imagem 1 não encontrada após 3 tentativas. Prosseguindo...")
    else:
        wb = load_workbook(caminho_do_arquivo)
        ws = wb[nome_da_aba] 
        if linha > ws.max_row:
            ws[coluna_ost + str(linha_especifica)] = 'CNH VENCIDA OU COM CARENCIA'
        else:
            ws[coluna_ost + str(linha_especifica)] = 'CNH VENCIDA OU COM CARENCIA'
        wb.save(caminho_do_arquivo)
        wb.close()

        while True:
            try:
                position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
                if position2:
                    center_x = position2.left + position2.width // 2
                    center_y = position2.top + position2.height // 2
                    pyautogui.click(center_x, center_y)
                    print("Imagem 2 foi encontrada na tela.")
                    break
                else:
                    print("Imagem 2 não encontrada na tela. Aguardando...")
            except Exception as e:
                print(f"Erro ao tentar encontrar a imagem 2: {e}")
            pyautogui.sleep(1)
        pyautogui.sleep(1)
        while True:
            try:
                position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
                if position3:
                    center_x = position3.left + position3.width // 2
                    center_y = position3.top + position3.height // 2
                    pyautogui.click(center_x, center_y)
                    print("Imagem 2 foi encontrada na tela.")
                    break
                else:
                    print("Imagem 2 não encontrada na tela. Aguardando...")
            except Exception as e:
                print(f"Erro ao tentar encontrar a imagem 2: {e}")
            pyautogui.sleep(1)
        pyautogui.sleep(1)
        while True:
            try:
                position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
                if position3:
                    center_x = position3.left + position3.width // 2
                    center_y = position3.top + position3.height // 2
                    pyautogui.click(center_x, center_y)
                    print("Imagem 2 foi encontrada na tela.")
                    break
                else:
                    print("Imagem 2 não encontrada na tela. Aguardando...")
            except Exception as e:
                print(f"Erro ao tentar encontrar a imagem 2: {e}")
            pyautogui.sleep(1)
        return aviso_ativo, aviso_veiculo  # Retorna o estado do aviso

aviso_ativo, aviso_veiculo = aviso_atencao('aviso_cnh_vencida.png','ok_marcado.png','ok_efetuado.png','curso_vencido.png')
