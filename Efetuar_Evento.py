import pandas as pd
import os
import numpy as np
import pyautogui
from datetime import datetime
import pyautogui
import pyperclip
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from unidecode import unidecode
import keyboard

caminho = os.getcwd() 
caminho_do_arquivo = 'EVENTO.xlsx'
nome_da_aba = 'Planilha1'
coluna_ost = 'J'


def click_image(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, "query_timeout_expered.png") 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Imagem de query_timeout_expered foi encontrada na tela.")
                click_image('ok.png')
                break
        except Exception as e:
            print("Aguardando...")
        pyautogui.sleep(1)

def liberado_efetuar(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Liberado para efetuar.")
                break
        except Exception as e:
            print("Não liberado para efetuar. Aguardando...")
        pyautogui.sleep(1)

def inclusao_documento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Tela de documento ainda nao fechou, aguarde.")  
        except Exception as e:
            print("Tela de documento fechada.")
            break
        pyautogui.sleep(1)

def click_aviso_veiculo(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position2:
                center_x = position2.left + position2.width // 2
                center_y = position2.top + position2.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def click(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            print("Imagem foi encontrada na tela.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def novo_evento(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Confirmação para inclusao de evento.")
                break
        except Exception as e:
            print("Não autorizado inclusão de evento.Aguardando...")
            try:
                position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
                if position2:
                    print("OK foi encontrada na tela.")
                    center_x = position.left + position.width // 2
                    center_y = position.top + position.height // 2
                    pyautogui.click(center_x, center_y)
                    click('ok_efetuado.png')
                    click('incluir.png')
                else:       
                    pyautogui.press('enter')
                    click('incluir.png')
            except Exception as e:
                print("OK não encontrada na tela.")

        pyautogui.sleep(1)

def confirmacao_documento_incluido(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Documento foi encontrada na tela.")
                break
        except Exception as e:
            print("Documento não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def numero_evento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Numero do evento nao foi encontrada na tela.Aguardando ...")
        except Exception as e:
            print("Numero do evento encontrado na tela.")
            break
        pyautogui.sleep(1)

def alerta_revisonais(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            print("Imagem foi encontrada na tela.")
            click_image('ok_marcado.png')
            wb = load_workbook(caminho_do_arquivo)
            ws = wb[nome_da_aba] 
            if linha > ws.max_row:
                ws[coluna_ost + str(linha_especifica)] = 'VEICULO COM ALERTAS DE REVISAO'
            else:
                ws[coluna_ost + str(linha_especifica)] = 'VEICULO COM ALERTAS DE REVISAO'
            wb.save(caminho_do_arquivo)
            wb.close()
            return True  # Indicar para pular a iteração
        else:
            print("Imagem não encontrada na tela.")
            return False
    except Exception as e:
        print(f"Erro ao tentar encontrar a imagem: {e}")
    try:
        position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
        if position2:
            print("Imagem foi encontrada na tela.")
            click_image('ok_marcado.png')
            wb = load_workbook(caminho_do_arquivo)
            ws = wb[nome_da_aba] 
            if linha > ws.max_row:
                ws[coluna_ost + str(linha_especifica)] = 'VEICULO COM ALERTAS DE REVISAO'
            else:
                ws[coluna_ost + str(linha_especifica)] = 'VEICULO COM ALERTAS DE REVISAO'
            wb.save(caminho_do_arquivo)
            wb.close()
            return True  # Indicar para pular a iteração
        else:
            print("Imagem não encontrada na tela.")
            return False
    except Exception as e:
        print(f"Erro ao tentar encontrar a imagem: {e}")
    return False
    pyautogui.sleep(1)

def motorista_nao_localizado(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            print("Imagem foi encontrada na tela.")
            wb = load_workbook(caminho_do_arquivo)
            ws = wb[nome_da_aba]
            if linha > ws.max_row:
                ws[coluna_ost + str(linha_especifica)] = 'MOTORISTA NAO ENCONTRADO'
            else:
                ws[coluna_ost + str(linha_especifica)] = 'MOTORISTA NAO ENCONTRADO'
            wb.save(caminho_do_arquivo)
            wb.close()
            return True  # Indicar para pular a iteração
        else:
            print("Imagem não encontrada na tela.")
            return False
    except Exception as e:
        print(f"Erro ao tentar encontrar a imagem: {e}")
        return False
    pyautogui.sleep(1)

def veiculo_nao_localizado(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            print("Imagem foi encontrada na tela.")
            wb = load_workbook(caminho_do_arquivo)
            ws = wb[nome_da_aba]
            if linha > ws.max_row:
                ws[coluna_ost + str(linha_especifica)] = 'VEICULO NAO ENCONTRADO OU INATIVO'
            else:
                ws[coluna_ost + str(linha_especifica)] = 'VEICULO NAO ENCONTRADO OU INATIVO'
            wb.save(caminho_do_arquivo)
            wb.close()
            return True  # Indicar para pular a iteração
        else:
            print("Imagem não encontrada na tela.")
            return False
    except Exception as e:
        print(f"Erro ao tentar encontrar a imagem: {e}")
        return False
    pyautogui.sleep(1)

def erro_rateio(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            print("Imagem foi encontrada na tela.")
            click_image('ok_efetuado.png')
            wb = load_workbook(caminho_do_arquivo)
            ws = wb[nome_da_aba] 
            if linha > ws.max_row:
                ws[coluna_ost + str(linha_especifica)] = 'ERRO DE RATEIO DO EVENTO'
            else:
                ws[coluna_ost + str(linha_especifica)] = 'ERRO DE RATEIO DO EVENTO'
            wb.save(caminho_do_arquivo)
            wb.close()
            return True  # Indicar para pular a iteração
        else:
            print("Imagem não encontrada na tela.")
            return False
    except Exception as e:
        print(f"Erro ao tentar encontrar a imagem: {e}")
        return False
    pyautogui.sleep(1)

def aviso_atencao(image_path, image_path2, image_path3, image_path4, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = 'IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    image_path4 = os.path.join(current_dir, caminho_imagem, image_path4)
    attempts = 3
    aviso_ativo = False  # Variável para indicar se o aviso está ativo
    aviso_veiculo = False  # Variável para indicar se o aviso do veículo está ativo

    # Tentar encontrar a primeira imagem até 3 vezes
    for attempt in range(attempts):
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                aviso_ativo = True
                break
            else:
                print(f"Imagem 1 não encontrada na tentativa {attempt + 1}. Aguardando...")
        except Exception as e:
            print(f"Erro ao tentar encontrar a imagem 1 na tentativa {attempt + 1}: {e}")
        
        try:
            position4 = pyautogui.locateOnScreen(image_path4, confidence=confidence)
            if position4:
                click_image('ok_marcado.png')
                aviso_veiculo = True
                break
        except Exception as e:
            print(f"Erro ao tentar encontrar a imagem 4 na tentativa {attempt + 1}: {e}")
              
        pyautogui.sleep(1)
        
    if not aviso_ativo:
        print("Imagem 1 não encontrada após 3 tentativas. Prosseguindo...")
    else:
        wb = load_workbook(caminho_do_arquivo)
        ws = wb[nome_da_aba]
        ws[coluna_ost + str(linha_especifica)] = 'CNH VENCIDA OU COM CARENCIA'
        wb.save(caminho_do_arquivo)
        wb.close()
        click_image('ok_marcado.png')
        pyautogui.sleep(0.5)
        click_image('ok_efetuado.png')
        pyautogui.sleep(0.5)
        pyautogui.press('enter')
        # click_image('ok_efetuado.png')
        pyautogui.sleep(0.5)

    
    return aviso_ativo, aviso_veiculo  # Retorna o estado do aviso

def aviso_antt(image_path,confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = 'IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    aviso_antt = False  # Variável para indicar se o aviso está ativo
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            aviso_antt = True  # Aviso encontrado, definir como ativo
        else:
            print(f"Imagem não encontrada na tentativa. Aguardando...")
    except Exception as e:
        print(f"Erro ao tentar encontrar a imagem: {e}")
    pyautogui.sleep(1)
    if not aviso_antt:
        print("Imagem  não encontrada. Prosseguindo...")
        return aviso_antt  # Retorna o estado do aviso
    wb = load_workbook(caminho_do_arquivo)
    ws = wb[nome_da_aba]  
    if linha > ws.max_row:
        ws[coluna_ost + str(linha_especifica)] = 'ANTT DO VEICULO ESTA VENCENDO'
    else:
        ws[coluna_ost + str(linha_especifica)] = 'ANTT DO VEICULO ESTA VENCENDO'
    wb.save(caminho_do_arquivo)
    wb.close()
    click_image('ok_marcado.png')
    pyautogui.sleep(2)
    return aviso_antt  # Retorna o estado do aviso

def erro_documento_naoencontrato(image_path, image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = 'IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    attempts = 3
    documento_nao_localizado = False  # Variável para indicar se o aviso está ativo
    # Tentar encontrar a primeira imagem até 3 vezes
    for attempt in range(attempts):
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                documento_nao_localizado = True  # Aviso encontrado, definir como ativo
                break
            else:
                print(f"Imagem 1 não encontrada na tentativa {attempt + 1}. Aguardando...")
        except Exception as e:
            print(f"Erro ao tentar encontrar a imagem 1 na tentativa {attempt + 1}: {e}")
        pyautogui.sleep(1)
    if not documento_nao_localizado:
        print("Imagem 1 não encontrada após 3 tentativas. Prosseguindo...")
        return documento_nao_localizado  # Retorna o estado do aviso
    wb = load_workbook(caminho_do_arquivo)
    ws = wb[nome_da_aba]  
    if linha > ws.max_row:
        ws[coluna_ost + str(linha_especifica)] = 'DOCUMENTO NAO ENCONTRADO'
    else:
        ws[coluna_ost + str(linha_especifica)] = 'DOCUMENTO NAO ENCONTRADO'
    wb.save(caminho_do_arquivo)
    wb.close()
    # Se a primeira imagem foi encontrada, tentar encontrar a segunda imagem
    while True:
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                center_x = position2.left + position2.width // 2
                center_y = position2.top + position2.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem 2 foi encontrada na tela.")
                break
            else:
                print("Imagem 2 não encontrada na tela. Aguardando...")
        except Exception as e:
            print(f"Erro ao tentar encontrar a imagem 2: {e}")
        pyautogui.sleep(1)
    pyautogui.sleep(0.5)
    click_image('cancelar_inclusao.png')
    return documento_nao_localizado  # Retorna o estado do aviso

def click_info_manifesto(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(40, 0)  # Movendo o cursor para cima
                pyautogui.click()  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

agora = datetime.now()
um_minuto_atras = agora - timedelta(minutes=1)
agora_formatado = agora.strftime("%d/%m/%Y %H:%M")
um_minuto_atras_formatado = um_minuto_atras.strftime("%H:%M")
#print(um_minuto_atras_formatado)

Planilha_eventos = pd.read_excel("EVENTO.xlsx")
Planilha_eventos['DATA'] = pd.to_datetime(Planilha_eventos['DATA'])
Planilha_eventos['DATA'] = Planilha_eventos['DATA'].dt.strftime('%d/%m/%Y')
# print(Planilha_eventos['DATA'])

linha_especifica = 1
pyautogui.sleep(2)

click_image('botao_frota.png')
click_image('frota_movimentacao.png')
click_image('frota_movimentacao_eventos.png')

for i, linha in enumerate(Planilha_eventos.index):
    data = Planilha_eventos.loc[linha, "DATA"]
    filial = Planilha_eventos.loc[linha, "FILIAL"]
    serie = Planilha_eventos.loc[linha, "SERIE"]
    manifesto = Planilha_eventos.loc[linha, "MANIFESTO"]
    codigo_evento = Planilha_eventos.loc[linha, "COD"]
    valor = Planilha_eventos.loc[linha, "VALOR"]        

    valor = str(valor)
    valor = valor.replace('.', ',')
    descricao = Planilha_eventos.loc[linha, "OBSERVAÇÃO"]
    descricao = str(descricao)
    if 'Ç' in descricao:
        descricao = descricao.replace("Ç", "C")
    elif('ç' in descricao):
        descricao = descricao.replace("ç", "c")
    descricao = unidecode(descricao)
    descricao = descricao.upper()
    placa = Planilha_eventos.loc[linha, "PLACA"]
    placa = str(placa)
    placa = placa.replace('*', '')
    placa = placa.replace('-', '')
    placa = placa.replace('_', '')
    placa = placa.replace('&', '')
    linha_especifica += 1
    click_image('incluir.png')
    pyautogui.sleep(2)
    novo_evento('campo_evento_automatico.png','ok_efetuado.png')
    click_info_manifesto('filial.png')
    for i in range(5):
        pyautogui.press('backspace')
    pyautogui.write(str(filial))
    pyautogui.press('tab')
    click_info_manifesto('data_evento.png')
    for i in range(10):
        pyautogui.press('backspace')
    for i in range(25):
        pyautogui.press('del')
    pyautogui.write(str(data))
    pyautogui.write(str(um_minuto_atras_formatado))
    for i in range(2):
        pyautogui.press('tab')
    pyautogui.sleep(0.2)
    click_info_manifesto('codigo_evento.png')
    pyautogui.write(str(codigo_evento))
    pyautogui.press('tab')
    click_info_manifesto('campo_veiculo.png') 
    pyautogui.sleep(0.5)  
    pyautogui.press('F2')
    click_info_manifesto('campo_placa.png')
    pyautogui.sleep(0.5) 
    pyautogui.write(str(placa))     
    click_image('situacao_veiculo.png')
    pyautogui.sleep(0.5)
    click_image('tipo_veiculo_normal.png')
    click_image('atualizar.png')
    pyautogui.sleep(2)
    click_image('selecionar.png')
    pyautogui.sleep(2)
    if veiculo_nao_localizado('campo_veiculo_vazio.png'):
        continue  
    for i in range(3):
        pyautogui.press('tab')
    pyautogui.press('enter')
    pyautogui.sleep(1)
    #se placa ou motorista der aviso colocar um ok 
    aviso_ativo, aviso_veiculo = aviso_atencao('aviso_cnh_vencida.png', 'ok_marcado.png', 'ok_efetuado.png', 'curso_vencido.png')
    alerta_revisonal = alerta_revisonais('ALERTA_REVISONAIS.png','aviso_dual_avencer.png')
    antt_vencida = aviso_antt('erro_antt.png')
    pyautogui.sleep(1)
    if motorista_nao_localizado('campo_motorista_vazio.png'):
        continue
    pyautogui.press('enter')  
    click_info_manifesto('campo_quantidade.png')
    pyautogui.write('1')
    pyautogui.press('tab')
    click_info_manifesto('campo_valor.png')
    pyautogui.write(valor)
    pyautogui.press('tab')
    click_info_manifesto('campo_observacao.png')
    pyautogui.sleep(3)
    pyautogui.write(descricao)
    pyautogui.sleep(2)
    click_image('salvar.png')
    numero_evento('campo_evento_automatico.png')
    click_info_manifesto('campo_evento.png')
    pyautogui.click(button='right')
    pyautogui.sleep(1)
    click_image('copy.png')
    pyautogui.sleep(0.5)
    try:
        text = pyperclip.paste()
        ost = int(text)
        print("Número do EVENTO:", ost)
    except ValueError:
        print("O conteúdo copiado não é um número válido.")
    except Exception as e:
        print("Ocorreu um erro:", str(e))

    wb = load_workbook(caminho_do_arquivo)
    ws = wb[nome_da_aba]
    coluna_evento = 'A'  
    if linha > ws.max_row:
        ws[coluna_evento + str(linha_especifica)] = ost
    else:
        ws[coluna_evento + str(linha_especifica)] = ost
    wb.save(caminho_do_arquivo)
    wb.close()

    click_image('inserir_manifesto.png')
    click_image('pasta_amarela.png')
    click_image('selecionar_tipo_documento.png')
    click_image('selecionar_manifesto.png')
    click_image('inserir_filial_manifesto.png')
    pyautogui.sleep(0.5)
    pyautogui.write(str(filial))
    pyautogui.sleep(0.5)
    pyautogui.press('tab')
    pyautogui.write(str(serie))
    pyautogui.sleep(0.5)
    pyautogui.press('tab')
    pyautogui.write(str(manifesto))
    pyautogui.sleep(0.5)
    pyautogui.press('tab')
    click_image('setinha_verde.png')
    pyautogui.sleep(2)
    confirmacao_documento_incluido('confirmacao_de_documento_incluido.png')
    #Verificar se o documento foi encontrado
    documento_nao_localizado = erro_documento_naoencontrato('erro_documento_naoencontrado.png','ok_efetuado.png')
    
    if documento_nao_localizado:
        print("Documento nao encontrado, ir para o proximo evento.")
        click_image('botao_voltar.png')
        pyautogui.sleep(1)
        pyautogui.press('tab')
        pyautogui.press('enter')
        inclusao_documento('tela_documento.png')        
        for i in range(2):
            pyautogui.press("tab")    
        pyautogui.press('enter')
        pyautogui.sleep(2)
        liberado_efetuar('botao_replicar_veiculo.png')
        print('Indo para o proximo evento')
    else:
        pyautogui.press('enter')
        click_image('botao_voltar.png')
        print("Documento encontrado. Prosseguir com outra ação.")
        if antt_vencida:
            click_image('ok_efetuado.png')
        if alerta_revisonal:
            click_image('ok_efetuado.png')
        if aviso_ativo:
            print("O aviso está ativo. Tome uma ação específica.")      

            click_image('ok_efetuado.png')
            pyautogui.sleep(1)
            click_image('ok_efetuado.png')
        else:
            print("O aviso não está ativo. Prosseguir com outra ação.")
        if aviso_veiculo:
            click_aviso_veiculo('ok_efetuado.png','yes_marcado.png')
        pyautogui.sleep(1)
        pyautogui.press('tab')
        pyautogui.press('enter')
        inclusao_documento('tela_documento.png')        
        for i in range(2):            pyautogui.press("tab")    
        pyautogui.press('enter')
        pyautogui.sleep(2)
        liberado_efetuar('botao_replicar_veiculo.png')
        click_image('efetuar.png')
        pyautogui.sleep(2)
        if erro_rateio('erro_rateio.png'):      
            continue  
        click_image('yes_efetuar.png')
        pyautogui.sleep(1)
        click_image('ok_efetuado.png')
        pyautogui.sleep(1)

pyautogui.sleep(2)
click_image('voltar.png')


        
