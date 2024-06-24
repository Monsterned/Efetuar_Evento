import pandas as pd
import os
import numpy as np
import pyautogui
from datetime import datetime
import pyautogui
import pyperclip
from datetime import datetime, timedelta
from openpyxl import load_workbook

caminho = os.getcwd() 

def click_image(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def click_info_manifesto(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(30, 0)  # Movendo o cursor para cima
                pyautogui.click()  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)



agora = datetime.now()
um_minuto_atras = agora - timedelta(minutes=1)
agora_formatado = agora.strftime("%d/%m/%Y %H:%M")
um_minuto_atras_formatado = um_minuto_atras.strftime("%H:%M")
#print(um_minuto_atras_formatado)

Planilha_eventos = pd.read_excel("EVENTO.xlsx")
Planilha_eventos['DATA'] = pd.to_datetime(Planilha_eventos['DATA'])
Planilha_eventos['DATA'] = Planilha_eventos['DATA'].dt.strftime('%d/%m/%Y')
# print(Planilha_eventos['DATA'])

linha_especifica = 2

pyautogui.sleep(2)
for i, linha in enumerate(Planilha_eventos.index):
    data = Planilha_eventos.loc[linha, "DATA"]
    filial = Planilha_eventos.loc[linha, "FILIAL"]
    serie = Planilha_eventos.loc[linha, "SERIE"]
    manifesto = Planilha_eventos.loc[linha, "MANIFESTO"]
    codigo_evento = Planilha_eventos.loc[linha, "COD"]
    valor = Planilha_eventos.loc[linha, "VALOR"]
    descricao = Planilha_eventos.loc[linha, "DESCRICAO"]
    placa = Planilha_eventos.loc[linha, "PLACA"]

    valor = str(valor)
    valor = valor.replace('.', ',')
    descricao = str(descricao)
    descricao = descricao.upper()
    print(descricao)
    pyautogui.write(descricao)

    # click_info_manifesto('campo_evento.png')
    # pyautogui.click(button='right')
    # for i in range(3):
    #     pyautogui.press("down")
    # pyautogui.sleep(1)
    # pyautogui.press("enter")

    # try:
    #     text = pyperclip.paste()
    #     ost = int(text)
    #     print("Número da OST:", ost)
    # except ValueError:
    #     print("O conteúdo copiado não é um número válido.")
    # except Exception as e:
    #     print("Ocorreu um erro:", str(e))

    # caminho_do_arquivo = 'EVENTO.xlsx'
    # nome_da_aba = 'Planilha1'
    # wb = load_workbook(caminho_do_arquivo)
    # ws = wb[nome_da_aba]
    # coluna_ost = 'A'  
    # if linha > ws.max_row:
    #     ws[coluna_ost + str(linha_especifica)] = ost
    # else:
    #     ws[coluna_ost + str(linha_especifica)] = ost
    # wb.save(caminho_do_arquivo)
    # wb.close()

    
    # pyautogui.write(str(data))
    # pyautogui.write(str(um_minuto_atras_formatado))

